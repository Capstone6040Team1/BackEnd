from flask import Flask, jsonify, request
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import OneHotEncoder, StandardScaler
import xgboost as xgb
from sklearn.metrics import mean_squared_error, r2_score

app = Flask(__name__)

# 文件路径（请修改为你的文件路径）
excel_file = r"./emp.xlsx"
workbook = load_workbook(excel_file)
sheet = workbook.active

# 读取 Excel 数据
def read_excel_data():
    """从 Excel 读取员工数据"""
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设第一行为表头
        data.append({
            'ID': row[0],
            'Name': row[1],
            'Job Role': row[2],
            'Job Level': row[3],
            'Total Experience': row[4],
            'Years in Current Role': row[5],
            'Education': row[6],
            'Department': row[7],
            'Previous Roles': row[8],
            'Projects Worked On': row[9],
            'Certifications': row[10],
            'Trainings Attended': row[11],
            'Performance Rating': row[12],
            'Skills': ', '.join(filter(None, [row[13], row[14], row[15]])),
            'Skill 1': row[13],
            'Skill 2': row[14],
            'Skill 3': row[15],
            'Skill 1 Frequency': row[16],
            'Skill 2 Frequency': row[17],
            'Skill 3 Frequency': row[18],
            'Skill 1 Importance': row[19],
            'Skill 2 Importance': row[20],
            'Skill 3 Importance': row[21]
        })
    return data

# 写入 Excel
def write_to_excel(data):
    """写入 Excel"""
    new_row = [
        data.get('ID'),
        data.get('Name'),
        data.get('Job Role'),
        data.get('Job Level'),
        data.get('Total Experience'),
        data.get('Years in Current Role'),
        data.get('Education'),
        data.get('Department'),
        data.get('Previous Roles'),
        data.get('Projects Worked On'),
        data.get('Certifications'),
        data.get('Trainings Attended'),
        data.get('Performance Rating'),
        data.get('Skill 1'),
        data.get('Skill 2'),
        data.get('Skill 3'),
        data.get('Skill 1 Frequency'),
        data.get('Skill 2 Frequency'),
        data.get('Skill 3 Frequency'),
        data.get('Skill 1 Importance'),
        data.get('Skill 2 Importance'),
        data.get('Skill 3 Importance')
    ]
    sheet.append(new_row)
    workbook.save(excel_file)

# 删除 Excel 行
def delete_excel_row(row_id):
    """删除 Excel 中某行"""
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == row_id:
            sheet.delete_rows(row[0].row)
            workbook.save(excel_file)
            return True
    return False

# 部门分类
def segregate_by_department():
    """按部门分类员工"""
    data = read_excel_data()
    department_dict = {}
    for employee in data:
        department = employee.get('Department', 'Unknown')  # 处理空值
        department_dict.setdefault(department, []).append(employee)
    return department_dict

# 训练 XGBoost 预测技能熟练度
def preprocess_and_train():
    """数据预处理并训练 XGBoost"""
    df = pd.DataFrame(read_excel_data())
    
    if df.empty:
        return None, None, None, None

    # One-Hot Encode Job Role
    ohe = OneHotEncoder(sparse_output=False, drop='first')
    encoded_roles = ohe.fit_transform(df[['Job Role']])
    df_encoded = pd.DataFrame(encoded_roles, columns=ohe.get_feature_names_out(['Job Role']))
    df = pd.concat([df, df_encoded], axis=1).drop(columns=['Job Role'])

    # 归一化
    scaler = StandardScaler()
    df[['Performance Rating']] *= 2  # 加权
    numeric_cols = ['Job Level', 'Total Experience', 'Projects Worked On', 'Certifications', 'Trainings Attended', 'Performance Rating']
    df[numeric_cols] = scaler.fit_transform(df[numeric_cols])

    # 计算技能得分
    for i in range(1, 4):
        df[f'Skill {i} Score'] = (
            0.15 * df[f'Skill {i} Frequency'] +
            0.15 * df[f'Skill {i} Importance'] +
            0.20 * df['Total Experience'] +
            0.05 * df['Projects Worked On'] +
            0.05 * df['Certifications'] +
            0.05 * df['Trainings Attended'] +
            0.35 * df['Performance Rating']
        )

    # 训练 XGBoost 模型
    features = df.drop(columns=[f'Skill {i}' for i in range(1, 4)] + [f'Skill {i} Frequency' for i in range(1, 4)] + [f'Skill {i} Importance' for i in range(1, 4)])

    models = {}
    for i in range(1, 4):
        model, _, _ = train_xgboost(features, df[f'Skill {i} Score'])
        models[f'Skill {i} Model'] = model

    return models, features

def train_xgboost(features, target):
    """训练 XGBoost 并计算 RMSE 和 R²"""
    X_train, X_test, y_train, y_test = train_test_split(features, target, test_size=0.2, random_state=42)
    model = xgb.XGBRegressor(objective='reg:squarederror', n_estimators=100, learning_rate=0.1, max_depth=5)
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)
    rmse = np.sqrt(mean_squared_error(y_test, y_pred))
    r2 = r2_score(y_test, y_pred)
    return model, rmse, r2

# API 端点
@app.route('/getAllData', methods=['GET'])
def get_all_data():
    return jsonify(read_excel_data())

@app.route('/getByDepartment', methods=['GET'])
def get_by_department():
    return jsonify(segregate_by_department())

@app.route('/addEmployee', methods=['POST'])
def add_employee():
    data = request.json
    if not data or 'ID' not in data or 'Name' not in data:
        return jsonify({'error': 'Invalid data'}), 400
    write_to_excel(data)
    return jsonify({'message': 'Employee added successfully'}), 200

@app.route('/predictSkillProficiency/<int:emp_id>', methods=['GET'])
def predict_skill_proficiency(emp_id):
    models, features = preprocess_and_train()
    employee_data = features[features['ID'] == emp_id]
    if employee_data.empty:
        return jsonify({"error": "Employee not found"}), 404
    return jsonify({f'Skill {i} Prediction': round(models[f'Skill {i} Model'].predict(employee_data)[0]) for i in range(1, 4)})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)
