from flask import Flask, jsonify, request
from openpyxl import load_workbook

app = Flask(__name__)

ML_API_URL = "http://your-ml-server.com/predict"

# Load the Excel file
excel_file = r"./emp.xlsx"
job_excel_file = r"./job_descriptions.xlsx"
workbook = load_workbook(excel_file)
sheet = workbook.active

# Helper functions to interact with the Excel file
def read_excel_data():
    """Read all rows of data from the Excel sheet."""
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True): # Assuming first row is header
        data.append({
            'id': row[0],
            'name': row[1],
            'job_title': row[2],
            'skills': row[3].split(',') if row[3] else [],
            'experience': row[4],
            'education': row[5],
            'department':row[6]
        })
    return data

def read_job_excel_data(file_path, sheet_name="Sheet1"):
    """Read all rows of job descriptions from the Excel sheet."""
    
    wb = load_workbook(filename=file_path, data_only=True)
    sheet = wb[sheet_name]
    
    job_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
        job_data.append({
            'job_id': row[0],
            'job_title': row[1],
            'required_skills': row[2].split(',') if row[2] else [],
            'experience': row[3],  # Expected experience in years
            'education': row[4],  # Required degree
            'job_level': row[5],  # Entry/Mid/Senior
            'certifications': row[6].split(',') if row[6] else []  # Required certifications
        })
    
    wb.close()
    return job_data

"""Job ID	Job Title	Required Skills	Experience	Education	  Job Level	  Certifications
   101	  Data Engineer	Python, SQL, AWS   3+ years	 Bachelor's	  Mid-Level	  AWS Certified
   102	  QA Engineer	 Selenium, Python  5+ years	 Master's	  Senior	  ISTQB"""

def write_to_excel(data):
    """Write a single row of data to the Excel sheet."""
    sheet.append(data)
    workbook.save(excel_file)

def update_excel_row(row_id, updated_data):
    """Update a specific row in the Excel sheet."""
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == row_id:  # Match by ID
            for key, value in updated_data.items():
                column_index = {'name': 2, 'job_title': 3, 'skills': 4, 'experience': 5, 'education': 6, 'department':7}
                if key in column_index:
                    row[column_index[key] - 1].value = ','.join(value) if key == 'skills' else value
            workbook.save(excel_file)
            return True
    return False

def delete_excel_row(row_id):
    """Delete a specific row from the Excel sheet."""
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == row_id:  # Match by ID
            sheet.delete_rows(row[0].row)
            workbook.save(excel_file)
            return True
    return False

def calculate_skill_metrics(data):
    """needed to be filled"""
    return {
        "skill_frequencies": {
            "Skill 1": {"skill": "Python", "frequency": 10},
            "Skill 2": {"skill": "AWS", "frequency": 8},
            "Skill 3": {"skill": "SQL", "frequency": 6}
        },
        "skill_importance": {
            "Skill 1": {"skill": "Python", "importance": 9},
            "Skill 2": {"skill": "AWS", "importance": 7},
            "Skill 3": {"skill": "SQL", "importance": 8}
        }
    }
    """return format"""

def calculate_job_skill(job_data):
    """Calculate job-related skill metrics from job descriptions."""
    return {
        "job_skill_frequencies": {
            "Skill 1": {"skill": "Python", "frequency": 12},
            "Skill 2": {"skill": "AWS", "frequency": 9},
            "Skill 3": {"skill": "SQL", "frequency": 7}
        },
        "job_skill_importance": {
            "Skill 1": {"skill": "Python", "importance": 10},
            "Skill 2": {"skill": "AWS", "importance": 8},
            "Skill 3": {"skill": "SQL", "importance": 7}
        }
    }

def segregate_by_department():
    """Segregate employees based on their department."""
    data = read_excel_data()
    department_dict = {}
    for employee in data:
        department = employee['department']
        if department not in department_dict:
            department_dict[department] = []
        department_dict[department].append(employee)
    return department_dict

# API Endpoints
@app.route('/getAllData', methods=['GET'])
def get_all_data():
    data = read_excel_data()
    return jsonify(data)

@app.route('/getByDepartment', methods=['GET'])
def get_by_department():
    data = segregate_by_department()
    return jsonify(data)

@app.route('/addEmployee', methods=['POST'])
def add_employee():
    data = request.json
    if not data or 'id' not in data or 'name' not in data:
        return jsonify({'error': 'Invalid data'}), 400

    new_row = [
        data['id'],
        data['name'],
        data.get('job_title', ''),
        ','.join(data.get('skills', [])),
        data.get('experience', ''),
        data.get('education', '')
    ]
    write_to_excel(new_row)
    return jsonify({'message': 'Employee added successfully'}), 200

@app.route('/updateEmployee/<int:emp_id>', methods=['POST'])
def update_employee(emp_id):
    data = request.json
    if not data:
        return jsonify({'error': 'Invalid data'}), 400

    if update_excel_row(emp_id, data):
        return jsonify({'message': 'Employee updated successfully'}), 200
    else:
        return jsonify({'error': 'Employee not found'}), 404

@app.route('/getSkillMetrics', methods=['GET'])
def get_skill_metrics():
    data = read_excel_data()
    skill_metrics = calculate_skill_metrics(data)
    return jsonify(skill_metrics)


@app.route('/deleteEmployee/<int:emp_id>', methods=['DELETE'])
def delete_employee(emp_id):
    if delete_excel_row(emp_id):
        return jsonify({'message': 'Employee deleted successfully'}), 200
    else:
        return jsonify({'error': 'Employee not found'}), 404

@app.route('/selfAssessment', methods=['POST'])
def self_assessment():
    data = request.json
    if not data or 'emp_id' not in data or 'assessment' not in data or 'score' not in data:
        return jsonify({'error': 'Invalid data'}), 400

    emp_id = data['emp_id']
    new_entry = f"{data['assessment']} (Score: {data['score']})"

    data_list = read_excel_data()
    employee = next((emp for emp in data_list if emp['id'] == emp_id), None)

    if employee:
        previous_assessment = employee.get('self_assessment', "")
        updated_assessment = previous_assessment + " | " + new_entry if previous_assessment else new_entry
        update_excel_row(emp_id, {'self_assessment': updated_assessment})
        return jsonify({'message': 'Self assessment updated successfully'}), 200
    return jsonify({'error': 'Employee not found'}), 404

@app.route('/hrAssessment', methods=['POST'])
def hr_assessment():
    data = request.json
    if not data or 'emp_id' not in data or 'assessment' not in data or 'score' not in data:
        return jsonify({'error': 'Invalid data'}), 400

    emp_id = data['emp_id']
    new_entry = f"{data['assessment']} (Score: {data['score']})"

    data_list = read_excel_data()
    employee = next((emp for emp in data_list if emp['id'] == emp_id), None)

    if employee:
        previous_assessment = employee.get('hr_assessment', "")
        updated_assessment = previous_assessment + " | " + new_entry if previous_assessment else new_entry
        update_excel_row(emp_id, {'hr_assessment': updated_assessment})
        return jsonify({'message': 'HR assessment updated successfully'}), 200
    return jsonify({'error': 'Employee not found'}), 404

@app.route('/getAssessment/<int:emp_id>', methods=['GET'])
def get_assessment(emp_id):
    data = read_excel_data()
    employee = next((emp for emp in data if emp['id'] == emp_id), None)

    if employee:
        return jsonify({
            "emp_id": emp_id,
            "self_assessment": employee.get('self_assessment', ""),
            "hr_assessment": employee.get('hr_assessment', "")
        }), 200
    return jsonify({"error": "Employee not found"}), 404

@app.route('/getEmployeeScore/<int:emp_id>', methods=['GET'])
def get_employee_score(emp_id):
    data_list = read_excel_data()
    employee = next((emp for emp in data_list if emp['id'] == emp_id), None)

    if not employee:
        return jsonify({"error": "Employee not found"}), 404
    input_data = {
        "skills": employee["skills"], 
        "experience": employee["experience"], 
        "education": employee["education"], 
        "self_assessment": employee["self_assessment"], 
        "hr_assessment": employee["hr_assessment"]
    }
    try:
        response = requests.post(ML_API_URL, json=input_data)
        ml_result = response.json()
        score = ml_result.get("score", None)
        if score is None:
            return jsonify({"error": "ML Model Error"}), 500

        return jsonify({
            "emp_id": emp_id,
            "ml_score": score
        }), 200
    except Exception as e:
        return jsonify({"error": f"ML Server Error: {str(e)}"}), 500

@app.route('/finalScore', methods=['GET'])
def final_score():
    emp_id = request.args.get('employee_id', type=int)
    job_id = request.args.get('job_id', type=int)

    # Read employee data
    employees = read_excel_data()
    employee = next((e for e in employees if e['id'] == emp_id), None)
    if not employee:
        return jsonify({"error": "Employee not found"}), 404

    # Read job description data
    job_descriptions = read_job_excel_data()
    job_description = next((j for j in job_descriptions if j['job_id'] == job_id), None)
    if not job_description:
        return jsonify({"error": "Job description not found"}), 404

    # Calculate employee skill score
    employee_skill_score = calculate_employee_skill_score(employee)

    # Calculate job skill score
    job_skill_score = calculate_job_skill_score(job_description)

    # Calculate final score
    if job_skill_score == 0:
        final_score = 0  # Avoid division by zero error
    else:
        final_score = employee_skill_score / job_skill_score

    return jsonify({
        "employee_id": emp_id,
        "job_id": job_id,
        "final_score": final_score
    }), 200

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)
